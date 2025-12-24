import openpyxl
import json
import os


class DataReader:
    """读取data目录下的两个文件"""

    def __init__(self):
        self.data_dir = "data"
        if not os.path.exists(self.data_dir):
            raise FileNotFoundError(f"data目录不存在：{self.data_dir}")

    def read_json(self):
        """读取test_data.json"""
        json_path = os.path.join(self.data_dir, "test_data.json")
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"test_data.json文件不存在：{json_path}")
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def read_excel(self, sheet_name="Login"):
        """读取test_data.xlsx,默认工作表Login"""
        excel_path = os.path.join(self.data_dir, "test_data.xlsx")
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"test_data.xlsx文件不存在：{excel_path}")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表{sheet_name}不存在")
        ws = wb[sheet_name]
        # 读取表头和数据
        headers = [cell.value for cell in ws[1] if cell.value]
        data = [
            dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)
        ]
        wb.close()
        return data
